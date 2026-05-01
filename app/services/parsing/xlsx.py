"""XLSX reader — each sheet rendered as a markdown table under an H2."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.services.parsing.base import Document, make_base_metadata


class XlsxReader:
    def load_data(self, file_path: str | Path) -> list[Document]:
        path = Path(file_path)
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)

        chunks: list[str] = []
        for sheet in wb.worksheets:
            rows = [
                [("" if c is None else str(c)).strip() for c in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            # drop fully-empty trailing rows / columns
            rows = [r for r in rows if any(c for c in r)]
            if not rows:
                continue
            # trim to longest non-empty column
            width = max(len(r) for r in rows)
            rows = [(r + [""] * (width - len(r)))[:width] for r in rows]

            chunks.append(f"## {sheet.title}")
            header = rows[0]
            chunks.append("| " + " | ".join(header) + " |")
            chunks.append("| " + " | ".join(["---"] * width) + " |")
            for r in rows[1:]:
                chunks.append("| " + " | ".join(c.replace("\n", " ") for c in r) + " |")
            chunks.append("")  # blank line between sheets

        if not chunks:
            return []
        text = "\n".join(chunks).rstrip()
        return [Document(
            text=text,
            metadata=make_base_metadata(file_path=path, parser="xlsx"),
        )]
